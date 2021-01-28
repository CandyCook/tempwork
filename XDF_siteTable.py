import xlwt
import xlrd
import openpyxl as xl
from openpyxl.styles import Font,Alignment
fontObj1 = Font(name=u'微软雅黑', bold=False, italic=False, size=12)
align = Alignment(horizontal='center', vertical='center')
#打开学生的名单
StudentFile = xl.load_workbook("学生表.xlsx")
StudentSheet = StudentFile[StudentFile.sheetnames[0]]
#打开目标座位表
sourceFile = xl.load_workbook("座位表.xlsx")
sourceSheet = sourceFile[sourceFile.sheetnames[0]]
#遍历学生名单
for row in range(3,StudentSheet.max_row +1 ):
    # 遍历目标座位表
    curStudentNum = StudentSheet.cell(row,4).value
    curStudentName = StudentSheet.cell(row, 2).value

    #在座位表中替换
    for i in sourceSheet.rows:
        for cell in i:
            if curStudentNum == cell.value:
                cell.value = curStudentName
            cell.font = fontObj1
#上面为正序座位表模块
print('已完成50%')
#当前总行数
CurRows = sourceSheet.max_row
#当前总列数
CurCols = sourceSheet.max_column

AllRows = CurRows * 2 +5
AllCols = CurCols +1


#以下为反向模块
for i in range(1,CurRows + 1):
    for j in range(1,CurCols + 1):
        sourceSheet.cell(AllRows-i,AllCols - j).value = sourceSheet.cell(i,j).value
        sourceSheet.cell(AllRows-i,AllCols - j).font = fontObj1
print('已完成95%')
#写入讲台位置
sourceSheet.merge_cells(start_row=CurRows+1,start_column=1,end_row=CurRows+1,end_column=CurCols)
sourceSheet.merge_cells(start_row=AllRows-CurRows-1,start_column=1,end_row=AllRows-CurRows-1,end_column=CurCols)

sourceSheet.cell(CurRows+1,1).value = '教室后部(助教视角)'
sourceSheet.cell(AllRows-CurRows-1,1).value = '教室后部(教师视角)'
sourceSheet.cell(CurRows+1,1).font = sourceSheet.cell(AllRows-CurRows-1,1).font = fontObj1
sourceSheet.cell(CurRows+1,1).alignment = sourceSheet.cell(AllRows-CurRows-1,1).alignment = align
sourceFile.save('排好后的座位表.xlsx')
print('完成！')


